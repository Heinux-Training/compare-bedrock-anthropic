#!/usr/bin/env python3
import boto3
import time
import statistics
import argparse
from datetime import datetime
import anthropic
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

def get_anthropic_api_key():
    """
    Get Anthropic API key from .env file.
    Raises ValueError if API key is not found.
    """
    api_key = os.getenv('ANTHROPIC_API_KEY')
    if not api_key:
        raise ValueError(
            "Anthropic API key not found in .env file. "
            "Please create a .env file with ANTHROPIC_API_KEY=your-api-key"
        )
    return api_key

def measure_direct_anthropic_latency(api_key, model_id, iterations=10):
    """
    Measure the latency of direct Anthropic API calls.
    """
    client = anthropic.Anthropic(api_key=api_key)
    
    latencies = []
    errors = 0
    
    print(f"Starting direct Anthropic API latency test")
    print(f"Model ID: {model_id}")
    print(f"Iterations: {iterations}")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print("-" * 50)
    
    prompt = "Write a paragraph starting with: 'Once upon a time...'"
    
    for i in range(iterations):
        try:
            start_time = time.time()
            
            response = client.messages.create(
                model=model_id,
                max_tokens=1000,
                messages=[
                    {"role": "user", "content": [{"type": "text", "text": prompt}]}
                ]
            )
            
            end_time = time.time()
            latency = (end_time - start_time) * 1000  # Convert to milliseconds
            latencies.append(latency)
            
            print(f"Request {i+1}: {latency:.2f} ms")
            
            # Add a small delay between requests
            time.sleep(1)
            
        except Exception as e:
            print(f"Error in request {i+1}: {str(e)}")
            errors += 1
    
    return calculate_statistics(latencies, errors)

def measure_bedrock_latency(source_region, target_region, model_id, iterations=10):
    """
    Measure the latency of Bedrock API calls from source_region to target_region.
    """
    # Create a Bedrock client in the target region
    bedrock_client = boto3.client('bedrock-runtime', region_name=target_region)
    
    latencies = []
    errors = 0
    
    print(f"Starting Bedrock latency test from {source_region} to {target_region}")
    print(f"Model ID: {model_id}")
    print(f"Iterations: {iterations}")
    print(f"Timestamp: {datetime.now().isoformat()}")
    print("-" * 50)
    
    # Simple prompt for testing
    prompt = "Write a paragraph starting with: 'Once upon a time...'"
    
    # Prepare request body based on model
    if "anthropic" in model_id:
        body = {
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 1000,
            "messages": [
                {"role": "user", "content": [{"type": "text", "text": prompt}]}
            ]
        }
    elif "amazon.titan" in model_id:
        body = {
            "inputText": prompt,
            "textGenerationConfig": {
                "maxTokenCount": 50
            }
        }
    else:
        print(f"Unsupported model: {model_id}")
        return
    
    import json
    body_str = json.dumps(body)
    
    for i in range(iterations):
        try:
            start_time = time.time()
            
            response = bedrock_client.invoke_model(
                modelId=model_id,
                body=body_str
            )
            
            end_time = time.time()
            latency = (end_time - start_time) * 1000  # Convert to milliseconds
            latencies.append(latency)
            
            print(f"Request {i+1}: {latency:.2f} ms")
            
            # Add a small delay between requests
            time.sleep(1)
            
        except Exception as e:
            print(f"Error in request {i+1}: {str(e)}")
            errors += 1
    
    return calculate_statistics(latencies, errors)

def calculate_statistics(latencies, errors):
    """
    Calculate and print statistics for latency measurements.
    """
    if latencies:
        avg_latency = statistics.mean(latencies)
        min_latency = min(latencies)
        max_latency = max(latencies)
        median_latency = statistics.median(latencies)
        p95_latency = sorted(latencies)[int(len(latencies) * 0.95)]
        
        print("\nResults:")
        print(f"Average latency: {avg_latency:.2f} ms")
        print(f"Minimum latency: {min_latency:.2f} ms")
        print(f"Maximum latency: {max_latency:.2f} ms")
        print(f"Median latency: {median_latency:.2f} ms")
        print(f"P95 latency: {p95_latency:.2f} ms")
        print(f"Successful requests: {len(latencies)}")
        print(f"Failed requests: {errors}")
        
        return {
            'avg_latency': avg_latency,
            'min_latency': min_latency,
            'max_latency': max_latency,
            'median_latency': median_latency,
            'p95_latency': p95_latency,
            'successful_requests': len(latencies),
            'failed_requests': errors
        }
    else:
        print("No successful requests to calculate latency.")
        return None

def save_to_excel(bedrock_stats, direct_stats, timestamp):
    """
    Save the benchmark results to an Excel file with formatted template.
    """
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Benchmark Results"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ["Metric", "Bedrock", "Direct API", "Difference", "Percentage Change"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # Data
        metrics = {
            'Average Latency (ms)': ('avg_latency', 'ms'),
            'Minimum Latency (ms)': ('min_latency', 'ms'),
            'Maximum Latency (ms)': ('max_latency', 'ms'),
            'Median Latency (ms)': ('median_latency', 'ms'),
            'P95 Latency (ms)': ('p95_latency', 'ms'),
            'Successful Requests': ('successful_requests', ''),
            'Failed Requests': ('failed_requests', '')
        }
        
        for row, (metric_name, (key, unit)) in enumerate(metrics.items(), 2):
            bedrock_value = bedrock_stats[key]
            direct_value = direct_stats[key]
            difference = bedrock_value - direct_value
            percentage = (difference / direct_value) * 100 if direct_value != 0 else float('inf')
            
            # Add data
            ws.cell(row=row, column=1, value=metric_name).border = border
            ws.cell(row=row, column=2, value=f"{bedrock_value:.2f}").border = border
            ws.cell(row=row, column=3, value=f"{direct_value:.2f}").border = border
            ws.cell(row=row, column=4, value=f"{difference:+.2f}").border = border
            ws.cell(row=row, column=5, value=f"{percentage:+.2f}%").border = border
        
        # Add metadata
        ws.cell(row=len(metrics) + 3, column=1, value="Test Information").font = Font(bold=True)
        ws.cell(row=len(metrics) + 4, column=1, value="Timestamp")
        ws.cell(row=len(metrics) + 4, column=2, value=timestamp)
        ws.cell(row=len(metrics) + 5, column=1, value="Total Requests")
        ws.cell(row=len(metrics) + 5, column=2, value=bedrock_stats['successful_requests'] + bedrock_stats['failed_requests'])
        
        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Create results directory if it doesn't exist
        results_dir = "benchmark_results"
        if not os.path.exists(results_dir):
            os.makedirs(results_dir)
        
        # Save the file
        filename = os.path.join(results_dir, f"benchmark_results_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(filename)
        print(f"\nResults saved to {filename}")
        
    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")
        print("Statistics data:")
        print("Bedrock stats:", bedrock_stats)
        print("Direct stats:", direct_stats)

def compare_methods(bedrock_stats, direct_stats):
    """
    Compare and print the differences between Bedrock and direct API calls.
    """
    if not bedrock_stats or not direct_stats:
        print("Cannot compare methods due to missing statistics.")
        return
    
    print("\nComparison between Bedrock and Direct Anthropic API:")
    print("-" * 50)
    
    metrics = {
        'Average Latency': ('avg_latency', 'ms'),
        'Minimum Latency': ('min_latency', 'ms'),
        'Maximum Latency': ('max_latency', 'ms'),
        'Median Latency': ('median_latency', 'ms'),
        'P95 Latency': ('p95_latency', 'ms'),
        'Successful Requests': ('successful_requests', ''),
        'Failed Requests': ('failed_requests', '')
    }
    
    for metric_name, (key, unit) in metrics.items():
        bedrock_value = bedrock_stats[key]
        direct_value = direct_stats[key]
        difference = bedrock_value - direct_value
        percentage = (difference / direct_value) * 100 if direct_value != 0 else float('inf')
        
        print(f"{metric_name}:")
        print(f"  Bedrock: {bedrock_value:.2f}{unit}")
        print(f"  Direct: {direct_value:.2f}{unit}")
        print(f"  Difference: {difference:+.2f}{unit} ({percentage:+.2f}%)")
        print()
    
    # Save results to Excel
    save_to_excel(bedrock_stats, direct_stats, datetime.now())

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Measure API latency for Bedrock and direct Anthropic API')
    parser.add_argument('--source-region', default='eu-north-1', help='Source region (where this script runs)')
    parser.add_argument('--target-region', default='us-east-1', help='Target region (where Bedrock API is called)')
    parser.add_argument('--model-id', default='arn:aws:bedrock:us-east-1:654654424505:inference-profile/us.anthropic.claude-sonnet-4-20250514-v1:0', help='Bedrock model ID to use')
    parser.add_argument('--iterations', type=int, default=10, help='Number of API calls to make')
    parser.add_argument('--anthropic-api-key', help='Anthropic API key for direct API calls (optional if set in .env file)')
    parser.add_argument('--direct-model-id', default='claude-sonnet-4-20250514', help='Model ID for direct Anthropic API calls')
    parser.add_argument('--compare', action='store_true', help='Compare Bedrock and direct API performance')
    
    args = parser.parse_args()
    
    if args.compare:
        # Use API key from command line if provided, otherwise from .env file
        api_key = args.anthropic_api_key or get_anthropic_api_key()
        bedrock_stats = measure_bedrock_latency(args.source_region, args.target_region, args.model_id, args.iterations)
        direct_stats = measure_direct_anthropic_latency(api_key, args.direct_model_id, args.iterations)
        compare_methods(bedrock_stats, direct_stats)
    else:
        bedrock_stats = measure_bedrock_latency(args.source_region, args.target_region, args.model_id, args.iterations) 